"""Build the spreadsheet report a non-technical reader can open without Python.

The findings in this project live in a notebook and a set of PNGs, which is fine for
an engineer and useless for anyone in a risk or finance team who works in Excel. This
writes the same aggregates -- the ones defined once in sql/aggregations.sql -- into a
formatted workbook with a pivot of fraud rate by channel and hour, so those numbers can
be sorted, filtered and charted by the people who actually act on them.

    python make_report.py            # reads app_data/, writes reports/fraud_summary.xlsx

Run `python load_sqlite.py` first if app_data/ is empty or stale.
"""

from __future__ import annotations

import os

import pandas as pd
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

AGG_DIR = "app_data"
OUT_DIR = "reports"
OUT = os.path.join(OUT_DIR, "fraud_summary.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="1C5CAB")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=13, bold=True)

#: sheet name -> (aggregate csv, one-line explanation for the reader)
SHEETS = {
    "Channel mix": ("channel_mix",
                    "Volume and fraud rate for every payment channel."),
    "Fraud by hour": ("fraud_by_hour",
                      "Fraud rate against hour of day. The strongest pattern in the data."),
    "Amount by class": ("amount_by_class",
                        "How far apart legitimate and fraudulent transactions sit on size."),
    "Balance signature": ("balance_signature",
                          "Money left the sender but the recipient balance never moved."),
    "Rule effectiveness": ("flagged_rule_effectiveness",
                           "What the platform's own built-in fraud rule actually catches."),
}


def load(name: str) -> pd.DataFrame:
    path = os.path.join(AGG_DIR, f"{name}.csv")
    if not os.path.exists(path):
        raise SystemExit(f"{path} not found. Run `python load_sqlite.py` first.")
    return pd.read_csv(path)


def build_pivot() -> pd.DataFrame:
    """Fraud rate as a channel x hour pivot table."""
    df = load("channel_hour")
    return df.pivot_table(index="hour", columns="channel",
                          values="fraud_rate_pct", aggfunc="sum").round(4)


def style_sheet(ws, note: str, n_cols: int) -> None:
    """Freeze and format the header, widen columns, and caption the sheet."""
    ws.insert_rows(1, 2)
    ws["A1"] = note
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(vertical="center")

    for col in range(1, n_cols + 1):
        cell = ws.cell(row=3, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        letter = get_column_letter(col)
        longest = max((len(str(c.value)) for c in ws[letter] if c.value is not None),
                      default=10)
        ws.column_dimensions[letter].width = min(max(longest + 2, 12), 34)

    ws.freeze_panes = "A4"
    # The filter must start at the header row, not at the caption above it.
    ws.auto_filter.ref = f"A3:{get_column_letter(n_cols)}{ws.max_row}"


def add_hour_chart(ws, n_rows: int) -> None:
    """A line chart of fraud rate by hour, so the shape is visible without a formula."""
    chart = LineChart()
    chart.title = "Fraud rate by hour of day"
    chart.y_axis.title = "Fraud rate (%)"
    chart.x_axis.title = "Hour"
    data = Reference(ws, min_col=4, min_row=3, max_row=3 + n_rows)
    cats = Reference(ws, min_col=1, min_row=4, max_row=3 + n_rows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height, chart.width = 8, 18
    ws.add_chart(chart, "G4")


def main() -> None:
    os.makedirs(OUT_DIR, exist_ok=True)

    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        for sheet, (name, note) in SHEETS.items():
            df = load(name)
            df.to_excel(writer, sheet_name=sheet, index=False)
            style_sheet(writer.sheets[sheet], note, len(df.columns))
            if name == "fraud_by_hour":
                add_hour_chart(writer.sheets[sheet], len(df))

        pivot = build_pivot()
        pivot.to_excel(writer, sheet_name="Pivot channel x hour")
        style_sheet(writer.sheets["Pivot channel x hour"],
                    "Fraud rate (%) by channel and hour of day.",
                    len(pivot.columns) + 1)

    print(f"wrote {OUT}")
    for sheet in list(SHEETS) + ["Pivot channel x hour"]:
        print(f"  sheet: {sheet}")


if __name__ == "__main__":
    main()
